# import os
# import sys
# sys.path.append(os.getcwd())

import os
from openpyxl import load_workbook
from utils.get_path import testdata_dir


class GetExcleData():
    def read_excle(self,excle_path):
        excle_path
        wb = load_workbook(excle_path)

    @classmethod
    def file_execute_list(cls):
        '''
        :return: 获取case下的所有用例文件列表,最多支持二级目录,通用排除文件返回最终要执行的用例文件
        '''
        excle_file_list = []
        # case_path = cls.project_directory + 'testdata'
        case_path = testdata_dir
        for filename in os.listdir(case_path):
            if 'xlsx' in filename:
                excle_file_list.append('case/' + filename)
            else:
                for i in os.listdir(case_path + '/' + filename):
                    if 'xlsx' in i:
                        excle_file_list.append('testdata/' + filename + '/' + i)

        return excle_file_list
    # def __init__(self,excle_path,sheet_name):
    #     # 1、加载一个excle 
    #     wb = load_workbook(excle_path)
    #     print(wb.sheetnames)            #获取所有工作表名称
    #     #  2、选择一个表单
    #     self.sh = wb[sheet_name]

    # def read_data(self):
    #     #获取表头信息
    #     keys = []
    #     for col_index in range(1,self.sh.max_column + 1) :
    #         keys.append(self.sh.cell(1, col_index).value)
    #     # print(keys)

    #     #获取每条测试用例的数据
    #     all_values = []
    #     for row_index in range(2, self.sh.max_row + 1) :
    #         all_data = []
            
    #         for col_index in range(1,self.sh.max_column + 1) :
    #             all_data.append(self.sh.cell(row_index,col_index).value)

    #             #赋予的这个值，只会给一次，需要创建一个列表，把值放入这个列表中，才能得到全部的数值
    #             case = dict(zip(keys,all_data))   
               
                     
    #         all_values.append(case)
    #     return all_values
        
if __name__ == "__main__" :
    #excle路径
    # excle_path = r"E:\InterfaceAuto\ApiAuto\testdata\login\ud3test.xlsx"
    # me = GetExcleData(excle_path,"用户登录")
    # cases = me.read_data()
    # for case in cases :
    # print(cases) 
    res = GetExcleData.file_execute_list()
    print(res)


            

